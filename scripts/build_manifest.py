"""
Build manifest.json from the curriculum spreadsheet.
Extracts Gutenberg IDs where possible, flags texts needing manual resolution.

Usage: python scripts/build_manifest.py [path_to_xlsx]
"""
import json
import os
import re
import sys
import openpyxl

DEFAULT_XLSX = "classical_curriculum_corpus.xlsx"

STAGE_MAP = {
    "0-Physical-World": {"stage": 0, "name": "physical-world"},
    "1-Foundation-Fables": {"stage": 1, "name": "foundation-fables"},
    "2-Grammar-Ancients": {"stage": 2, "name": "grammar-ancients"},
    "3-Logic-Medieval-Ren": {"stage": 3, "name": "logic-medieval-renaissance"},
    "4-Rhetoric-Modern": {"stage": 4, "name": "rhetoric-modern"},
    "5-Science-NatPhil": {"stage": 5, "name": "science-natural-philosophy"},
    "6-Drama-Poetry": {"stage": 6, "name": "drama-poetry"},
}

def extract_gutenberg_id(url_str):
    if not url_str or url_str in ("N/A", "Partial"):
        return None
    m = re.search(r'gutenberg\.org/ebooks/(\d+)', url_str)
    if m:
        return int(m.group(1))
    return None

def build_manifest(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    manifest = {
        "version": 1,
        "description": "Classical curriculum training corpus manifest",
        "texts": [],
        "stats": {"total": 0, "gutenberg_auto": 0, "manual_needed": 0, "excluded": 0},
    }

    for sheet_name, stage_info in STAGE_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for row in rows:
            if not row[1]:
                continue

            title = row[1]
            pd_status = row[5] or ""

            if pd_status == "No":
                manifest["stats"]["excluded"] += 1
                continue

            # Cast seq to int: openpyxl may return float (1.0) which
            # survives JSON round-trip and breaks f-string :02d formatting
            raw_seq = row[0]
            seq = int(raw_seq) if raw_seq is not None else 0

            gid = extract_gutenberg_id(row[7])
            slug = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')[:60]

            entry = {
                "stage": stage_info["stage"],
                "stage_name": stage_info["name"],
                "seq": seq,
                "title": title,
                "author": row[2] or "",
                "date": str(row[3]) if row[3] else "",
                "slug": slug,
                "gutenberg_id": gid,
                "source_url": row[7] or "",
                "pd_status": pd_status,
                "translation": row[6] or "",
                "token_est_k": row[8] or "",
                "notes": row[10] or "",
                "fetch_status": "auto" if gid else "manual",
            }
            manifest["texts"].append(entry)
            manifest["stats"]["total"] += 1
            if gid:
                manifest["stats"]["gutenberg_auto"] += 1
            else:
                manifest["stats"]["manual_needed"] += 1

    wb.close()
    return manifest

if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

    os.makedirs("config", exist_ok=True)

    manifest = build_manifest(xlsx)

    out_path = "config/manifest.json"
    with open(out_path, "w") as f:
        json.dump(manifest, f, indent=2)

    print(f"Manifest: {manifest['stats']['total']} texts")
    print(f"  Auto-fetchable (Gutenberg): {manifest['stats']['gutenberg_auto']}")
    print(f"  Manual resolution needed:   {manifest['stats']['manual_needed']}")
    print(f"  Excluded (not PD):          {manifest['stats']['excluded']}")
    print()
    if manifest["stats"]["manual_needed"] > 0:
        print("Texts needing manual source resolution:")
        for t in manifest["texts"]:
            if t["fetch_status"] == "manual":
                print(f"  Stage {t['stage']}: {t['title']}")
                print(f"    URL hint: {t['source_url']}")
